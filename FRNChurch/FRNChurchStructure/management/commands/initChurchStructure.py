from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from FRNChurchStructure.models import Group

class Command(BaseCommand):
    help = 'initChurchStructure: This command will init the Church Structure from the \
    		xlsx file /static/FRNChurchStructure/InputData/ChurchStructure.xlsx'

    def __init__(self):
    	self.groupType = []

    def handle(self, *args, **options):
    	wb = load_workbook('FRNChurchStructure/static/FRNChurchStructure/InputData/ChurchStructure.xlsx')

    	activeSheet = wb.active
    	columnCount = activeSheet.max_column
    	rowCount = activeSheet.max_row
    	for column in range(1, columnCount+1):
    		cell = activeSheet.cell(row=1, column=column)
    		self.groupType.append(cell.value)

    	print(self.groupType)

    	for rowindex in range(2, rowCount+1):
    		row = activeSheet[rowindex]
    		self.processRow(row)

    def processRow(self, row):
    	parent = None
    	cellCount = len(row)
    	for cellIndex in range(0, cellCount):
    		group = self.getGroup(row[cellIndex].value, self.groupType[cellIndex], parent)
    		parent = group

    def getGroup(self, groupName, type, parent):
    	try:
    		group = Group.objects.get(name=groupName, type = type)
    	except Group.DoesNotExist:
    		group = Group()
    		group.name = groupName
    		group.type = type
    		if None != parent:
    			group.parentGroup = parent
    		group.save()
    		print('new %s \'%s\' created under %s' %(type, groupName, parent))
    	return group
