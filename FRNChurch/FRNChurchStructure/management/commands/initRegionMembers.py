from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from FRNChurchStructure.models import Group, Christian

class Command(BaseCommand):
    help = 'initRegionMembers: This command will init the members of a Region from the \
    		xlsx file /static/FRNChurchStructure/InputData/RegionMembers.xlsx'

    def __init__(self):
    	self.zoneIndex = -1
    	self.fgIndex = -1
    	self.nameIndex = -1
    	self.genderIndex = -1
    	self.dobirthIndex = -1
    	self.dobapIndex = -1
    	self.phoneIndex = -1
    	self.emailIndex = -1

    def handle(self, *args, **options):
    	wb = load_workbook('FRNChurchStructure/static/FRNChurchStructure/InputData/RegionMembers.xlsx')

    	activeSheet = wb.active
    	columnCount = activeSheet.max_column
    	rowCount = activeSheet.max_row
    	for column in range(1, columnCount+1):
    		cell = activeSheet.cell(row=1, column=column)
    		if cell.value == "Zone":
    			self.zoneIndex = column-1
    		if cell.value == "Family Group":
    			self.fgIndex = column-1
    		if cell.value == "Gender":
    			self.genderIndex = column-1
    		if cell.value == "Name":
    			self.nameIndex = column-1
    		if cell.value == "Phone":
    			self.phoneIndex = column-1
    		if cell.value == "Email":
    			self.emailIndex = column-1
    		if cell.value == "DateOfBirth":
    			self.dobirthIndex = column-1
    		if cell.value == "DateOfBaptism":
    			self.dobapIndex = column-1

    	if self.zoneIndex == -1:
    		print("Zone column not present")
    		return
    	if self.fgIndex == -1:
    		print("Family Group column not present")
    		return
    	if self.nameIndex == -1:
    		print("Name column not present")
    		return
    	if self.genderIndex == -1:
    		print("Gender column not present")
    		return
    	if self.dobirthIndex == -1:
    		print("DateOfBirth column not present")
    		return
    	if self.dobapIndex == -1:
    		print("DateOfBaptism column not present")
    		return
    	if self.phoneIndex == -1:
    		print("Phone column not present")
    		return
    	if self.emailIndex == -1:
    		print("Email column not present")
    		return

    	for rowindex in range(2, rowCount+1):    
    		row = activeSheet[rowindex]
    		if False == self.processRow(row):
    			print("Error: Row not processed: ", rowindex)

    def processRow(self, row):
    	zoneName = row[self.zoneIndex].value
    	fgName = row[self.fgIndex].value
    	gender = row[self.genderIndex].value
    	name = row[self.nameIndex].value
    	dobrith = row[self.dobirthIndex].value
    	dobap = row[self.dobapIndex].value
    	phone = row[self.phoneIndex].value
    	email = row[self.emailIndex].value

    	try:
    	    fg = Group.objects.get(name = fgName, type = "Family Group")
    	except Group.DoesNotExist:
    		print( "Family group doesnot exist: ", fgName)
    		return False

    	try:
    	    zone = Group.objects.get(name = zoneName, type = "Zone")
    	except Group.DoesNotExist:
    		print( "Zone doesnot exist: ", zoneName)
    		return False

    	if zone.id != fg.parentGroup.id:
    		print("Family group not in zone")
    		print("Family group: ", fgName)
    		print("Zone: ", fg.parentGroup_id)
    		return False

    	return True




