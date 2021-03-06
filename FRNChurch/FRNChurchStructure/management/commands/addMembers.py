from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from FRNChurchStructure.models import Group, Christian, Contact

class Command(BaseCommand):
    help = 'addMembers: This command will add members to the Church groups.\
            New groups will be created if the specified group doesnot exist'

    def add_arguments(self, parser):
        parser.add_argument('xlsFilePath', type=str, help='Indicates the name of the file with member info')

    def __init__(self):
        self.regionName = 'KAT'

    def handle(self, *args, **options):
        wb = load_workbook(options['xlsFilePath'])
        activeSheet = wb.active

        if not self.isValid(activeSheet[1]):
            return

        try:
            rgn = Group.objects.get(name=self.regionName)
        except Group.DoesNotExist:
            print("Error: Region not found ", self.regionName)
            print("Error: Create region without type manually and then try addMembers")
            return

        rowCount = activeSheet.max_row
        for rowindex in range(2, rowCount+1):
            self.processRow(activeSheet[rowindex])
 
    def processRow(self, row):
        chrch = self.getChurch(row[0].value)
        region = self.getGroupUnder(row[1].value, 'Region', chrch)
        sector = self.getGroupUnder(row[2].value, 'Sector', region)
        zone = self.getGroupUnder(row[3].value, 'Zone', sector)
        fg = self.getGroupUnder(row[4].value, 'Family Group', zone)

        newCtn = self.createChristian(fg, row[5].value, row[6].value, row[7].value,
            row[8].value, row[9].value)
        self.updateContact(newCtn, row[10].value, row[11].value)

    def updateContact(self, ctn, phone, email):
        if phone == None and email == None:
            return
        contact = Contact()
        if phone != None:
            contact.phoneNumber = phone
        if email != None:
            contact.email = email
        contact.save()
        ctn.contact = contact
        ctn.save()

    def createChristian(self, fg, name, gender, dobirth, dobaptism, mstatus):
        chrtn = Christian()
        chrtn.familyGroup = fg
        names = name.split(" ", 1)
        chrtn.firstName = names[0]
        if len(names) == 2:
            chrtn.secondName = names[1]
        chrtn.gender = gender
        chrtn.dateOfBirth = dobirth
        chrtn.dateOfBaptism = dobaptism
        if "Married" in mstatus or "Parent" in mstatus:
            chrtn.maritalStatus = 'M'
        if "Widow" in mstatus:
            chrtn.maritalStatus = 'W'
        if "Divor" in mstatus:
            chrtn.maritalStatus = 'D'
        chrtn.save()
        print('New Christian %s \'%s\' created and added under %s %s'
            %(chrtn.firstName, chrtn.gender, fg.type, fg.name))
        return chrtn

    def getChurch(self, chrchName):
        try:
            region = Group.objects.get(name=self.regionName)
        except Group.DoesNotExist:
            print("Error: Region not found ", self.regionName)
            return

        chrch = self.getGroupUnder(chrchName, 'Church', region)
        return chrch

    def getGroupUnder(self, name, type, parent):
        try:
            gp = Group.objects.get(name=name, type=type, parentGroup=parent)
            return gp
        except Group.DoesNotExist:
            gp = Group()
            gp.name = name
            gp.type = type
            gp.parentGroup = parent
            gp.save()
            print('New %s \'%s\' created under %s %s' %(gp.type, gp.name, parent.type, parent))
            return gp

    def isValid(self, title):
        if "Church" != title[0].value:
            print("Error: Invalid xls file: First colum should have Church")
            return False
        if "Region" != title[1].value:
            print("Error: Invalid xls file: Second colum should have Region")
            return False
        if "Sector" != title[2].value:
            print("Error: Invalid xls file: Third colum should have Sector")
            return False
        if "Zone" != title[3].value:
            print("Error: Invalid xls file: Fourth colum should have Zone")
            return False
        if "Family Group" != title[4].value:
            print("Error: Invalid xls file: Fifth colum should have Family Group")
            return False
        if "Name" != title[5].value:
            print("Error: Invalid xls file: Sixth colum should have Name")
            return False
        if "Gender" != title[6].value:
            print("Error: Invalid xls file: Seventh colum should have Gender")
            return False
        if "DOBirth" != title[7].value:
            print("Error: Invalid xls file: Seventh colum should have DOBirth")
            return False
        if "DOBaptism" != title[8].value:
            print("Error: Invalid xls file: Seventh colum should have DOBaptism")
            return False
        if "Marital Status" != title[9].value:
            print("Error: Invalid xls file: Seventh colum should have Marital Status")
            return False
        if "Phone Number" != title[10].value:
            print("Error: Invalid xls file: Seventh colum should have Phone Number")
            return False
        if "Email" != title[11].value:
            print("Error: Invalid xls file: Seventh colum should have Email")
            return False

        return True



